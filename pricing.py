"""Pricing engine — all prices from CSV / quote workbook, never guessed.

Sources (in priority order per line item):
  1. Job folder: Purchased Components Quote.csv (Module6121 output)
  2. Job folder: Pullcore Prices.csv (Module6121 output)
  3. Quote / steel Excel workbook (BMS #2 steel block + summary)
  4. Shop: Purchased Components Prices.csv (same file the macro reads)
  5. Dimensions: quote/steel Excel workbook when present, else CAD export

Parts & Pricing tab sections mirror the quote workbook:
  - Steel Plates / Mold Base
  - Pull Cores & Keys
  - Purchased Components
"""
from __future__ import annotations

import json
from pathlib import Path

from . import config, sheet_pricing
from .roles import role_group, role_label


def load_rates() -> dict:
    """Legacy Settings endpoint — returns shop CSV snapshot for display only."""
    shop = sheet_pricing.load_shop_prices()
    rates = {}
    for row in shop:
        comp = (row.get("Component") or "").strip()
        price = sheet_pricing._safe_float(row.get("UnitPrice"))
        if comp and price > 0:
            rates[comp] = {"mode": "flat", "rate": price, "minimum": 0.0, "source": "csv"}
    if config.PRICING_CONFIG_PATH.exists():
        try:
            saved = json.loads(config.PRICING_CONFIG_PATH.read_text(encoding="utf-8"))
            rates.update(saved)
        except Exception:
            pass
    return rates


def save_rates(rates: dict) -> dict:
    config.PRICING_CONFIG_PATH.write_text(json.dumps(rates, indent=2), encoding="utf-8")
    return load_rates()


def _job_dir(job_id: str) -> Path:
    safe = job_id.strip().replace("..", "").replace("/", "_")
    return config.JOBS_ROOT / safe


def _line_from_section_row(
    row: dict,
    *,
    index: str,
    section: str,
    role_group_name: str,
) -> dict:
    role = row.get("role") or ""
    return {
        "index": index,
        "section": section,
        "component": row.get("component") or "",
        "role": role,
        "role_label": role_label(role) if role else (row.get("component") or ""),
        "role_group": role_group_name,
        "confidence": "HIGH",
        "quote": True,
        "price": float(row.get("price") or 0),
        "price_source": row.get("price_source") or "",
        "thickness": row.get("thickness"),
        "width": row.get("width"),
        "length": row.get("length"),
        "qty": row.get("qty"),
        "cu_in": row.get("cu_in"),
        "hours": row.get("hours"),
        "vendor": row.get("vendor") or "",
        "part_number": row.get("part_number") or "",
        "unit_price": row.get("unit_price"),
        "material": row.get("material") or "",
        "category": row.get("category") or "",
    }


def build_quote_sheet(job: dict) -> dict:
    job_id = job.get("job_id", "")
    job_dir = _job_dir(job_id)
    base_type = (job.get("base_type") or "").lower()

    shop_rows = sheet_pricing.load_shop_prices()
    job_purchased = sheet_pricing.load_job_purchased_quote(job_dir)
    sheet_dims = sheet_pricing.read_sheet_dimensions(job_dir)

    steel_rows = sheet_pricing.load_steel_plate_lines(job_dir)
    pullcore_rows = sheet_pricing.load_pullcore_lines(job_dir)
    purchased_rows = sheet_pricing.load_purchased_lines(job_dir)
    summary = sheet_pricing.load_quote_summary(job_dir)

    steel_items = [
        _line_from_section_row(
            r,
            index=f"S{i}",
            section="steel",
            role_group_name="Steel Plates / Mold Base",
        )
        for i, r in enumerate(steel_rows, start=1)
    ]
    pullcore_items = [
        _line_from_section_row(
            r,
            index=f"K{i}",
            section="pullcore",
            role_group_name="Pull Cores & Keys",
        )
        for i, r in enumerate(pullcore_rows, start=1)
    ]
    purchased_items = [
        _line_from_section_row(
            r,
            index=f"P{i}",
            section="purchased",
            role_group_name="Purchased Components",
        )
        for i, r in enumerate(purchased_rows, start=1)
    ]

    line_items: list[dict] = []
    total = 0.0
    csv_priced = 0
    missing = 0

    # Classified CAD parts (standard mold bases). Skip on BMS — steel/pullcore/purchased
    # sections are the source of truth and avoid empty AI A/B/rail rows.
    classified_items: list[dict] = []
    if base_type != "bms":
        for row in job.get("parts", []):
            priced = sheet_pricing.price_for_part(row, shop_rows, job_purchased, sheet_dims)
            price = priced["price"]
            total += price
            if price > 0:
                csv_priced += 1
            elif row.get("quote") or row.get("Quote"):
                missing += 1

            role = row.get("role") or ""
            item = {
                "index": row.get("index"),
                "section": "classified",
                "component": row.get("Component") or row.get("component"),
                "role": role,
                "role_label": row.get("role_label") or role_label(role),
                "role_group": row.get("role_group") or role_group(role),
                "confidence": row.get("confidence") or row.get("Confidence"),
                "quote": bool(row.get("quote") or row.get("Quote")),
                "price": price,
                "price_source": priced.get("price_source", ""),
                "thickness": priced.get("thickness"),
                "width": priced.get("width"),
                "length": priced.get("length"),
                "qty": sheet_pricing._safe_float(row.get("Qty") or row.get("QTY") or 1, 1.0),
            }
            classified_items.append(item)
            line_items.append(item)

    # Macro sections — always surface when present (BMS and standard).
    for item in steel_items + pullcore_items + purchased_items:
        price = float(item.get("price") or 0)
        total += price
        if price > 0:
            csv_priced += 1
        else:
            # Steel hours/price may be Excel-formula-only (data_only needs a prior Excel save).
            # Don't count steel blanks as "missing CSV price".
            if item.get("section") != "steel":
                missing += 1
        line_items.append(item)

    # Prefer workbook grand total when available (includes machining + commission).
    display_total = total
    if summary.get("grand_total_finish"):
        display_total = float(summary["grand_total_finish"])
    elif summary.get("grand_total_rough"):
        display_total = float(summary["grand_total_rough"])
    elif summary.get("total_price_finish"):
        display_total = float(summary["total_price_finish"])
    elif summary.get("total_price_rough"):
        display_total = float(summary["total_price_rough"])

    sources = []
    if steel_items:
        sources.append("quote/steel workbook")
    if pullcore_items:
        sources.append("Pullcore Prices.csv")
    if purchased_items:
        sources.append("Purchased Components Quote.csv")
    if classified_items:
        sources.append("Purchased Components Prices.csv")

    return {
        "job_id": job_id,
        "line_items": line_items,
        "sections": {
            "steel": steel_items,
            "pullcore": pullcore_items,
            "purchased": purchased_items,
            "classified": classified_items,
        },
        "steel_plates": steel_items,
        "pullcore_components": pullcore_items,
        "purchased_components": purchased_items,
        "summary": summary,
        "total_price": round(display_total, 2),
        "section_total_price": round(total, 2),
        "quoted_part_count": sum(1 for li in line_items if li.get("quote")),
        "total_part_count": len(line_items),
        "csv_priced_count": csv_priced,
        "missing_csv_price_count": missing,
        "pricing_source": " + ".join(sources) if sources else "Purchased Components Prices.csv",
        "shop_csv": str(config.PURCHASED_PRICES_CSV),
        "has_steel_sheet_dims": bool(sheet_dims) or bool(steel_items),
    }

# CMS FINAL PRICING OUTPUT PATCH V4
# Appended at end of pricing.py intentionally.
#
# Final layer cleanup before browser sees pricing rows.
#
# Rules:
#   1. Do not read quote / purchased files until Module6121 is DONE.
#   2. For BMS jobs, replace stale "--" / "sheet" steel rows with rows from
#      the final Quote Steel Grinding workbook.
#   3. Quote Steel Grinding workbook columns:
#        A = Description
#        C = Qty
#        D = Thickness
#        E = Width
#        F = Length
#        H = Price
#   4. Purchased Components Quote.csv is loaded only after Module6121 is DONE.

_CMS_OUT_BMS_ORDER = [
    "TCP",
    "BCP",
    "ID HOLDER",
    "OD HOLDER",
    "ID POT BLOCK",
    "OD POT BLOCK",
]

_CMS_OUT_BMS_ALIASES = {
    "TCP": "TCP",
    "TOP CLAMP": "TCP",
    "TOP CLAMP PLATE": "TCP",
    "TOP CLAMPING": "TCP",
    "TOP CLAMPING PLATE": "TCP",
    "TOP SMED": "TCP",
    "ID SMED": "TCP",

    "BCP": "BCP",
    "BOTTOM CLAMP": "BCP",
    "BOT CLAMP": "BCP",
    "BOTTOM CLAMP PLATE": "BCP",
    "BOT CLAMP PLATE": "BCP",
    "BOTTOM CLAMPING": "BCP",
    "BOTTOM CLAMPING PLATE": "BCP",
    "BOTTOM SMED": "BCP",
    "BOT SMED": "BCP",
    "OD SMED": "BCP",

    "ID HOLDER": "ID HOLDER",
    "ID HOLDER BLOCK": "ID HOLDER",
    "ID HOLDER MATERIAL": "ID HOLDER",
    "TOP HOLDER": "ID HOLDER",
    "TOP HOLDER BLOCK": "ID HOLDER",
    "ID MOLD BASE": "ID HOLDER",
    "ID MOLDBASE": "ID HOLDER",
    "TOP MOLD BASE": "ID HOLDER",
    "TOP MOLDBASE": "ID HOLDER",

    "OD HOLDER": "OD HOLDER",
    "OD HOLDER BLOCK": "OD HOLDER",
    "OD HOLDER MATERIAL": "OD HOLDER",
    "BOTTOM HOLDER": "OD HOLDER",
    "BOT HOLDER": "OD HOLDER",
    "BOTTOM HOLDER BLOCK": "OD HOLDER",
    "BOT HOLDER BLOCK": "OD HOLDER",
    "OD MOLD BASE": "OD HOLDER",
    "OD MOLDBASE": "OD HOLDER",
    "BOTTOM MOLD BASE": "OD HOLDER",
    "BOT MOLD BASE": "OD HOLDER",

    "ID POT": "ID POT BLOCK",
    "ID POT BLOCK": "ID POT BLOCK",
    "ID POT BLOCK MATERIAL": "ID POT BLOCK",
    "TOP POT": "ID POT BLOCK",
    "TOP POT BLOCK": "ID POT BLOCK",
    "TCP POT": "ID POT BLOCK",
    "TCP POT BLOCK": "ID POT BLOCK",

    "OD POT": "OD POT BLOCK",
    "OD POT BLOCK": "OD POT BLOCK",
    "OD POT BLOCK MATERIAL": "OD POT BLOCK",
    "BOTTOM POT": "OD POT BLOCK",
    "BOT POT": "OD POT BLOCK",
    "BOTTOM POT BLOCK": "OD POT BLOCK",
    "BOT POT BLOCK": "OD POT BLOCK",
    "BCP POT": "OD POT BLOCK",
    "BCP POT BLOCK": "OD POT BLOCK",
}


def _cms_out_norm(value):
    return " ".join(
        str(value or "")
        .replace("_", " ")
        .replace("-", " ")
        .replace(".", " ")
        .upper()
        .split()
    )


def _cms_out_key(value):
    return (
        str(value or "")
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
        .replace(".", "")
        .replace("#", "")
    )


def _cms_out_float(value, default=0.0):
    try:
        s = str(value or "").strip().replace("$", "").replace(",", "")
        if not s:
            return default
        return float(s)
    except Exception:
        return default


def _cms_out_qty(value, default=1):
    try:
        s = str(value or "").strip()
        if not s:
            return default
        n = int(float(s))
        return n if n > 0 else default
    except Exception:
        return default


def _cms_out_role(value):
    n = _cms_out_norm(value)

    if n in _CMS_OUT_BMS_ALIASES:
        return _CMS_OUT_BMS_ALIASES[n]

    for key, role in _CMS_OUT_BMS_ALIASES.items():
        if key in n:
            return role

    return None


def _cms_out_kv(path):
    data = {}
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                data[k.strip()] = v.strip()
    except Exception:
        pass
    return data


def _cms_out_cnum_from_text(value):
    import re

    m = re.search(r"\bC\d{4,8}\b", str(value or ""), re.IGNORECASE)
    if not m:
        return ""
    return m.group(0).upper()


def _cms_out_find_cnum_in_obj(obj):
    if obj is None:
        return ""

    if isinstance(obj, (str, int, float)):
        return _cms_out_cnum_from_text(obj)

    if isinstance(obj, dict):
        for k, v in obj.items():
            if str(k).lower() in ("cnum", "c_number", "job_id", "jobid", "id", "name", "folder", "path"):
                c = _cms_out_cnum_from_text(v)
                if c:
                    return c

        for v in obj.values():
            c = _cms_out_find_cnum_in_obj(v)
            if c:
                return c

    if isinstance(obj, (list, tuple)):
        for v in obj:
            c = _cms_out_find_cnum_in_obj(v)
            if c:
                return c

    return ""


def _cms_out_job_dir(args, kwargs, result=None):
    from pathlib import Path

    for key in ("job_dir", "job_folder", "folder", "path", "job_path", "job_id", "job"):
        if key in kwargs:
            c = _cms_out_cnum_from_text(kwargs[key])
            if c:
                return Path(r"C:\CMS_Local_Workspace") / c

            try:
                p = Path(kwargs[key])
                c = _cms_out_cnum_from_text(str(p))
                if c:
                    return Path(r"C:\CMS_Local_Workspace") / c
                return p
            except Exception:
                pass

    for arg in args:
        c = _cms_out_cnum_from_text(arg)
        if c:
            return Path(r"C:\CMS_Local_Workspace") / c

        try:
            p = Path(arg)
            c = _cms_out_cnum_from_text(str(p))
            if c:
                return Path(r"C:\CMS_Local_Workspace") / c
            if str(p):
                return p
        except Exception:
            pass

    c = _cms_out_find_cnum_in_obj(result)
    if c:
        return Path(r"C:\CMS_Local_Workspace") / c

    return None


def _cms_out_macro_done(job_dir):
    from pathlib import Path

    if job_dir is None:
        return False

    job_dir = Path(job_dir)

    status = _cms_out_kv(Path(r"C:\CMS_Local_Workspace") / "cms_macro_status.txt")
    st = str(status.get("Status", "")).upper()
    folder = str(status.get("CurrentJobFolder", "")).strip().lower()

    if st == "STARTED" and folder and folder == str(job_dir).strip().lower():
        return False

    if st in ("DONE", "COMPLETED") and folder and folder == str(job_dir).strip().lower():
        return True

    log_path = job_dir / "CMS_Base_Export_Log.txt"
    try:
        if log_path.exists():
            txt = log_path.read_text(encoding="utf-8-sig", errors="replace").upper()
            if "DONE JOB" in txt or "DONE ACTIVE CAD QUOTE" in txt or "RUNACTIVEASSEMBLY COMPLETED" in txt:
                return True
    except Exception:
        pass

    return False


def _cms_out_find_quote_workbooks(job_dir):
    from pathlib import Path

    if job_dir is None:
        return []

    job_dir = Path(job_dir)
    matches = []

    try:
        for p in job_dir.rglob("*"):
            if not p.is_file():
                continue

            name = p.name
            low = name.lower()

            if name.startswith("~$"):
                continue

            if not (low.endswith(".xlsx") or low.endswith(".xlsm") or low.endswith(".xls")):
                continue

            if "quote" in low and "steel" in low and "grind" in low:
                matches.append(p)
    except Exception:
        pass

    matches.sort(key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)
    return matches


def _cms_out_make_steel_row(role, qty, thickness, width, length, price, source, row_num):
    return {
        "description": role,
        "Description": role,
        "name": role,
        "Name": role,
        "label": role,
        "display_name": role,
        "part_name": role,
        "component": role,
        "role": role,
        "resolved_name": role,

        "role_group_name": "Steel Plates / Mold Base",
        "group": "Steel Plates / Mold Base",
        "category": "Steel Plates / Mold Base",

        "qty": qty,
        "QTY": qty,
        "quantity": qty,

        "thickness": thickness,
        "Thickness": thickness,
        "width": width,
        "Width": width,
        "length": length,
        "Length": length,

        "hours": None,
        "Hours": None,

        "price": price,
        "Price": price,
        "total": price,
        "Total": price,
        "extended": price,
        "Extended": price,
        "unit_price": price,
        "UnitPrice": price,
        "price_usd": price,
        "amount": price,
        "cost": price,

        "quoted": True,
        "quote": True,

        # Important: do not put "sheet" here. Frontend showed that in Price.
        "pricing_note": "",
        "price_note": "",

        "source": source,
        "price_source": source,
        "workbook_row": row_num,
    }


def _cms_out_score(qty, thickness, width, length, price, row_num):
    score = row_num
    if qty > 0:
        score += 1000
    if thickness > 0 and width > 0 and length > 0:
        score += 1000
    if price > 0:
        score += 1000
    return score


def _cms_out_rows_openpyxl(wb_path):
    try:
        import openpyxl
    except Exception:
        return []

    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    except Exception:
        return []

    if "QuoteWorksheet" in wb.sheetnames:
        ws = wb["QuoteWorksheet"]
    elif "Quote" in wb.sheetnames:
        ws = wb["Quote"]
    else:
        ws = wb[wb.sheetnames[0]]

    best = {}

    for r in range(1, (ws.max_row or 0) + 1):
        label = ws.cell(r, 1).value
        role = _cms_out_role(label)

        if not role:
            continue

        qty = _cms_out_qty(ws.cell(r, 3).value, 0)
        thickness = _cms_out_float(ws.cell(r, 4).value)
        width = _cms_out_float(ws.cell(r, 5).value)
        length = _cms_out_float(ws.cell(r, 6).value)
        weight = _cms_out_float(ws.cell(r, 7).value)
        price = _cms_out_float(ws.cell(r, 8).value)

        if price <= 0:
            pp = _cms_out_float(ws.cell(r, 2).value)
            if pp > 0 and weight > 0:
                price = pp * weight

        if qty <= 0 or thickness <= 0 or width <= 0 or length <= 0:
            continue

        score = _cms_out_score(qty, thickness, width, length, price, r)

        if role not in best or score >= best[role]["_score"]:
            row = _cms_out_make_steel_row(
                role=role,
                qty=qty,
                thickness=thickness,
                width=width,
                length=length,
                price=price,
                source="workbook:" + str(wb_path),
                row_num=r,
            )
            row["_score"] = score
            best[role] = row

    ordered = []
    for role in _CMS_OUT_BMS_ORDER:
        if role in best:
            row = dict(best[role])
            row.pop("_score", None)
            ordered.append(row)

    return ordered


def _cms_out_rows_excel_com(wb_path):
    try:
        import win32com.client
    except Exception:
        return []

    xl = None
    wb = None

    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False

        wb = xl.Workbooks.Open(str(wb_path), ReadOnly=True)

        ws = None
        for sheet in wb.Worksheets:
            if str(sheet.Name).lower() in ("quoteworksheet", "quote"):
                ws = sheet
                break

        if ws is None:
            ws = wb.Worksheets(1)

        used_rows = int(ws.UsedRange.Rows.Count)
        best = {}

        for r in range(1, used_rows + 1):
            label = ws.Cells(r, 1).Value
            role = _cms_out_role(label)

            if not role:
                continue

            qty = _cms_out_qty(ws.Cells(r, 3).Value, 0)
            thickness = _cms_out_float(ws.Cells(r, 4).Value)
            width = _cms_out_float(ws.Cells(r, 5).Value)
            length = _cms_out_float(ws.Cells(r, 6).Value)
            weight = _cms_out_float(ws.Cells(r, 7).Value)
            price = _cms_out_float(ws.Cells(r, 8).Value)

            if price <= 0:
                pp = _cms_out_float(ws.Cells(r, 2).Value)
                if pp > 0 and weight > 0:
                    price = pp * weight

            if qty <= 0 or thickness <= 0 or width <= 0 or length <= 0:
                continue

            score = _cms_out_score(qty, thickness, width, length, price, r)

            if role not in best or score >= best[role]["_score"]:
                row = _cms_out_make_steel_row(
                    role=role,
                    qty=qty,
                    thickness=thickness,
                    width=width,
                    length=length,
                    price=price,
                    source="workbook:" + str(wb_path),
                    row_num=r,
                )
                row["_score"] = score
                best[role] = row

        ordered = []
        for role in _CMS_OUT_BMS_ORDER:
            if role in best:
                row = dict(best[role])
                row.pop("_score", None)
                ordered.append(row)

        return ordered

    except Exception:
        return []

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass

        try:
            if xl is not None:
                xl.Quit()
        except Exception:
            pass


def _cms_out_final_steel_rows(job_dir):
    workbooks = _cms_out_find_quote_workbooks(job_dir)

    for wb_path in workbooks:
        rows = []

        if str(wb_path).lower().endswith((".xlsx", ".xlsm")):
            rows = _cms_out_rows_openpyxl(wb_path)

        if len(rows) < 4:
            rows = _cms_out_rows_excel_com(wb_path)

        if len(rows) >= 4:
            return rows

    return []


def _cms_out_pc_get(row, *names, default=""):
    lookup = {}
    for k, v in (row or {}).items():
        lookup[_cms_out_key(k)] = v

    for name in names:
        k = _cms_out_key(name)
        if k in lookup:
            return lookup[k]

    return default


def _cms_out_find_purchased_csv(job_dir):
    from pathlib import Path

    if job_dir is None:
        return None

    job_dir = Path(job_dir)

    candidates = [
        job_dir / "Purchased Components Quote.csv",
        job_dir / "documents" / "Purchased Components Quote.csv",
        job_dir / "Purchased Components.csv",
        job_dir / "documents" / "Purchased Components.csv",
    ]

    for p in candidates:
        if p.exists():
            return p

    try:
        for p in job_dir.rglob("Purchased Components Quote.csv"):
            if p.is_file():
                return p
    except Exception:
        pass

    return None


def _cms_out_purchased_rows(job_dir):
    import csv

    csv_path = _cms_out_find_purchased_csv(job_dir)

    if csv_path is None:
        return []

    rows = []

    with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.DictReader(f)

        for raw in reader:
            component = " ".join(str(_cms_out_pc_get(raw, "Component", "Comp", "Name", "Item") or "").split())
            vendor = " ".join(str(_cms_out_pc_get(raw, "Vendor", "Manufacturer", "MFG") or "").split())
            part_number = " ".join(str(_cms_out_pc_get(raw, "PartNumber", "Part Number", "PartNo", "Part No", "Part #") or "").split())
            description = " ".join(str(_cms_out_pc_get(raw, "Description", "Desc") or "").split())

            qty = _cms_out_qty(_cms_out_pc_get(raw, "QTY", "Qty", "Quantity"), 1)
            unit_price = _cms_out_float(_cms_out_pc_get(raw, "UnitPrice", "Unit Price", "Price Each"), 0.0)
            extended = _cms_out_float(_cms_out_pc_get(raw, "Extended", "Ext", "Total", "Extended Price"), 0.0)

            if extended <= 0 and unit_price > 0:
                extended = unit_price * qty

            label = component or description or part_number

            if not label:
                continue

            if _cms_out_key(label) == "total":
                continue

            rows.append({
                "description": label,
                "Description": label,
                "name": label,
                "Name": label,
                "label": label,
                "display_name": label,
                "role": "purchased_component",
                "resolved_name": label,
                "role_group_name": "Purchased Components",
                "group": "Purchased Components",
                "category": "Purchased Components",
                "vendor": vendor,
                "part_number": part_number,
                "partNo": part_number,
                "part_no": part_number,
                "component": component,
                "component_description": description,
                "qty": qty,
                "QTY": qty,
                "quantity": qty,
                "unit_price": unit_price,
                "UnitPrice": unit_price,
                "price": extended,
                "Price": extended,
                "total": extended,
                "Total": extended,
                "extended": extended,
                "quoted": True,
                "quote": True,
                "pricing_note": "",
                "source": "Purchased Components Quote.csv",
                "price_source": "job_csv:Purchased Components Quote.csv",
                "csv_path": str(csv_path),
            })

    return rows


def _cms_out_row_is_steel(row):
    if not isinstance(row, dict):
        return False

    for key in ("role_group_name", "group", "category"):
        if _cms_out_norm(row.get(key)) == "STEEL PLATES / MOLD BASE":
            return True

    role = _cms_out_role(row.get("role") or row.get("description") or row.get("name") or row.get("label"))
    if role in _CMS_OUT_BMS_ORDER:
        return True

    desc = str(row.get("description") or row.get("name") or row.get("label") or "").strip()
    if desc in ("--", ""):
        t = _cms_out_float(row.get("thickness") or row.get("Thickness"))
        w = _cms_out_float(row.get("width") or row.get("Width"))
        l = _cms_out_float(row.get("length") or row.get("Length"))
        if t > 0 and w > 0 and l > 0:
            return True

    return False


def _cms_out_row_is_purchased(row):
    if not isinstance(row, dict):
        return False

    for key in ("role_group_name", "group", "category"):
        if _cms_out_norm(row.get(key)) == "PURCHASED COMPONENTS":
            return True

    if _cms_out_norm(row.get("role")) == "PURCHASED COMPONENT":
        return True

    return False


def _cms_out_postprocess(obj, steel_rows, purchased_rows):
    if isinstance(obj, list):
        has_steel = any(_cms_out_row_is_steel(x) for x in obj if isinstance(x, dict))
        has_purchased = any(_cms_out_row_is_purchased(x) for x in obj if isinstance(x, dict))

        new_list = list(obj)

        if has_steel and steel_rows:
            new_list = [x for x in new_list if not (isinstance(x, dict) and _cms_out_row_is_steel(x))]
            new_list = steel_rows + new_list

        if has_purchased and purchased_rows:
            new_list = [x for x in new_list if not (isinstance(x, dict) and _cms_out_row_is_purchased(x))]
            new_list = new_list + purchased_rows

        return [_cms_out_postprocess(x, steel_rows, purchased_rows) for x in new_list]

    if isinstance(obj, dict):
        copied = {}

        group_name = _cms_out_norm(
            obj.get("role_group_name")
            or obj.get("group")
            or obj.get("category")
            or obj.get("name")
            or obj.get("title")
        )

        for k, v in obj.items():
            if group_name == "STEEL PLATES / MOLD BASE" and k in ("items", "rows", "line_items", "lineItems", "parts"):
                copied[k] = steel_rows if steel_rows else v
            elif group_name == "PURCHASED COMPONENTS" and k in ("items", "rows", "line_items", "lineItems", "parts"):
                copied[k] = purchased_rows if purchased_rows else v
            else:
                copied[k] = _cms_out_postprocess(v, steel_rows, purchased_rows)

        if group_name == "STEEL PLATES / MOLD BASE" and steel_rows:
            for ck in ("count", "line_items", "lineItemsCount", "qty"):
                if ck in copied and not isinstance(copied.get(ck), list):
                    copied[ck] = len(steel_rows)

        if group_name == "PURCHASED COMPONENTS" and purchased_rows:
            for ck in ("count", "line_items", "lineItemsCount", "qty"):
                if ck in copied and not isinstance(copied.get(ck), list):
                    copied[ck] = len(purchased_rows)

        return copied

    return obj


def _cms_out_apply_patch():
    import inspect

    g = globals()

    for name, obj in list(g.items()):
        if name.startswith("_cms_out_"):
            continue

        if not inspect.isfunction(obj):
            continue

        if getattr(obj, "__module__", None) != __name__:
            continue

        if getattr(obj, "_cms_out_wrapped", False):
            continue

        def make_wrapper(fn):
            def wrapper(*args, **kwargs):
                result = fn(*args, **kwargs)
                job_dir = _cms_out_job_dir(args, kwargs, result)

                # Absolute rule: while macro is running, remove early pricing rows.
                if not _cms_out_macro_done(job_dir):
                    return _cms_out_postprocess(result, [], [])

                steel_rows = _cms_out_final_steel_rows(job_dir)
                purchased_rows = _cms_out_purchased_rows(job_dir)

                return _cms_out_postprocess(result, steel_rows, purchased_rows)

            wrapper.__name__ = getattr(fn, "__name__", "wrapper")
            wrapper.__doc__ = getattr(fn, "__doc__", None)
            wrapper._cms_out_wrapped = True
            return wrapper

        g[name] = make_wrapper(obj)


_cms_out_apply_patch()
